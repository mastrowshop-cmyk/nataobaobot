from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, CallbackQueryHandler, filters
)
from openpyxl import load_workbook

from config import BOT_TOKEN, ADMIN_ID, WEBAPP_URL
from database import SessionLocal
from models import User, Parcel, Settings

# ---------- START ----------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()
    user = db.query(User).filter(User.tg_id == update.effective_user.id).first()

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton(
            "📦 Открыть приложение",
            web_app=WebAppInfo(url=WEBAPP_URL)
        )
    ]])

    if not user:
        context.user_data["register"] = True
        await update.message.reply_text(
            "Добро пожаловать в Nataobao 👋\n\n"
            "Введите:\nКОД Имя Фамилия\n\n"
            "Пример:\nNTB123 Иван Иванов",
            reply_markup=keyboard
        )
        return

    if user.role == "pending":
        await update.message.reply_text(
            "⏳ Ваша заявка на рассмотрении",
            reply_markup=keyboard
        )
        return

    await update.message.reply_text(
        f"Добро пожаловать, {user.name}!",
        reply_markup=keyboard
    )

# ---------- CALLBACK ----------
async def callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()
    query = update.callback_query
    await query.answer()

    action, tg_id = query.data.split("_")
    user = db.query(User).filter(User.tg_id == int(tg_id)).first()

    if action == "ok":
        user.role = "client"
        db.commit()
        await context.bot.send_message(user.tg_id, "✅ Заявка одобрена")
    else:
        db.delete(user)
        db.commit()
        await context.bot.send_message(user.tg_id, "❌ Заявка отклонена")

# ---------- TEXT ----------
async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()

    # регистрация
    if context.user_data.get("register"):
        try:
            code, name, surname = update.message.text.split(maxsplit=2)
        except:
            await update.message.reply_text("❌ Формат: КОД Имя Фамилия")
            return

        user = User(
            tg_id=update.effective_user.id,
            code=code,
            name=f"{name} {surname}",
            role="pending"
        )
        db.add(user)
        db.commit()

        await context.bot.send_message(
            ADMIN_ID,
            f"🔔 Новая заявка\n{user.name}\nКод: {code}",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Одобрить", callback_data=f"ok_{user.tg_id}"),
                InlineKeyboardButton("❌ Отклонить", callback_data=f"no_{user.tg_id}")
            ]])
        )

        context.user_data.clear()
        await update.message.reply_text("⏳ Заявка отправлена админу")

# ---------- МОИ ДОСТАВКИ ----------
async def my(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()
    user = db.query(User).filter(User.tg_id == update.effective_user.id).first()
    parcels = db.query(Parcel).filter(Parcel.user_code == user.code).all()

    if not parcels:
        await update.message.reply_text("📦 У вас нет доставок")
        return

    for p in parcels:
        await update.message.reply_text(
            f"📦 {p.description}\n"
            f"Статус: {p.status}\n"
            f"Сумма: {p.price or '-'} ₽"
        )

# ---------- ВЗВЕШИВАНИЕ ----------
async def weigh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()
    user = db.query(User).filter(User.tg_id == update.effective_user.id).first()
    if user.role != "operator":
        return
    context.user_data["weigh"] = True
    await update.message.reply_text("Введите: КОД ВЕС")

# ---------- IMPORT EXCEL ----------
async def import_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    db = SessionLocal()
    user = db.query(User).filter(User.tg_id == update.effective_user.id).first()
    if user.role != "admin":
        return

    file = await update.message.document.get_file()
    await file.download_to_drive("import.xlsx")

    wb = load_workbook("import.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code, desc = row
        db.add(Parcel(user_code=code, description=desc))

    db.commit()
    await update.message.reply_text("✅ Excel импортирован")

# ---------- RUN ----------
def run_bot():
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("my", my))
    app.add_handler(CommandHandler("weigh", weigh))
    app.add_handler(CommandHandler("import", import_excel))
    app.add_handler(CallbackQueryHandler(callback))
    app.add_handler(MessageHandler(filters.TEXT, text_handler))
    app.run_polling()
